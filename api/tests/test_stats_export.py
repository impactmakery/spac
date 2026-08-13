"""The usage page as a workbook.

The point of building it here rather than in the browser is that the charts
are real ones — Excel chart XML pointing at ranges on the sheets — so these
tests care that the charts exist and refer to the data, not merely that a file
came back.
"""

from io import BytesIO

import pytest


@pytest.fixture()
def world(db):
    from datetime import date, datetime, timedelta

    from app.core.security import hash_password
    from app.models import DailyMetric, Municipality, User
    from app.services.metrics import TZ

    pw = hash_password("stats-export-1")
    one = Municipality(name="City One")
    two = Municipality(name="עיריית שתיים")
    db.add_all([one, two])
    db.flush()
    db.add_all(
        [
            User(email="root@x.org", role="system_admin", status="active",
                 password_hash=pw, name="Root"),
            User(email="admin@x.org", role="municipality_admin", municipality=one,
                 status="active", password_hash=pw, name="Admin"),
            User(email="worker@x.org", role="department_user", municipality=one,
                 status="active", password_hash=pw, name="Worker"),
        ]
    )

    today = datetime.now(TZ).date()
    for offset in range(3):
        day = today - timedelta(days=offset)
        db.add(DailyMetric(day=day, active_users=3 + offset, chat_sessions=2,
                           chat_messages=5, unanswered=1, board_items=2,
                           comments=1, likes=1, files_uploaded=4))
        for muni in (one, two):
            db.add(DailyMetric(day=day, municipality_id=muni.id, active_users=1,
                               chat_sessions=1, chat_messages=2, unanswered=1,
                               board_items=1, comments=0, likes=0, files_uploaded=2))
    db.commit()
    assert isinstance(today, date)
    return {"one": one, "two": two}


def auth(client, email):
    r = client.post("/api/auth/login", json={"email": email, "password": "stats-export-1"})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def load(content: bytes):
    from openpyxl import load_workbook

    return load_workbook(BytesIO(content))


def test_the_workbook_has_a_sheet_per_shape_of_data(client, world):
    """A summary, a series, a comparison and a list do not share columns, so
    they do not share a sheet."""
    r = client.get("/api/stats/platform.xlsx?range_days=7&lang=he",
                   headers=auth(client, "root@x.org"))
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]

    wb = load(r.content)
    assert wb.sheetnames == ["סיכום", "לאורך זמן", "לפי רשות"]


def test_the_charts_are_real_ones_pointing_at_the_data(client, world):
    """Not pictures. A number changed in Excel has to move the chart, which is
    the whole reason this is not a screenshot."""
    r = client.get("/api/stats/platform.xlsx?range_days=7",
                   headers=auth(client, "root@x.org"))
    wb = load(r.content)

    over_time = wb["לאורך זמן"]
    assert len(over_time._charts) == 2
    line = over_time._charts[0]
    ref = line.series[0].val.numRef.f
    assert "לאורך זמן" in ref  # the range is on this sheet, by name

    comparison = wb["לפי רשות"]
    assert len(comparison._charts) == 1
    assert comparison._charts[0].type == "bar"


def test_the_numbers_match_the_page(client, world):
    body = client.get("/api/stats/platform?range_days=7",
                      headers=auth(client, "root@x.org")).json()
    r = client.get("/api/stats/platform.xlsx?range_days=7",
                   headers=auth(client, "root@x.org"))
    wb = load(r.content)

    summary = {row[0].value: row[1].value for row in wb["סיכום"].iter_rows(min_row=5)}
    assert summary["משתמשים פעילים"] == body["kpis"]["active_users"]
    assert summary["שאלות שנשאלו"] == body["kpis"]["chat_messages"]


def test_a_hebrew_workbook_opens_right_to_left(client, world):
    """Excel moves column A to the right and flips the chart axes; faking it
    with alignment would leave the charts pointing the wrong way."""
    r = client.get("/api/stats/platform.xlsx?lang=he&range_days=7",
                   headers=auth(client, "root@x.org"))
    assert load(r.content)["סיכום"].sheet_view.rightToLeft is True


def test_an_english_workbook_does_not(client, world):
    r = client.get("/api/stats/platform.xlsx?lang=en&range_days=7",
                   headers=auth(client, "root@x.org"))
    wb = load(r.content)
    assert wb.sheetnames[0] == "Summary"
    assert wb["Summary"].sheet_view.rightToLeft is False


def test_the_filename_survives_being_downloaded(client, world):
    r = client.get("/api/stats/platform.xlsx?range_days=7",
                   headers=auth(client, "root@x.org"))
    disposition = r.headers["content-disposition"]
    assert "attachment" in disposition
    assert "usage-platform-7d.xlsx" in disposition


@pytest.mark.parametrize("email", ["admin@x.org", "worker@x.org"])
def test_only_a_system_admin_gets_the_platform_workbook(client, world, email):
    """It holds every municipality's figures, so it stays with the role that
    already sees them all."""
    r = client.get("/api/stats/platform.xlsx", headers=auth(client, email))
    assert r.status_code == 404


def test_a_municipality_admin_gets_their_own(client, world):
    r = client.get("/api/stats/municipality.xlsx?range_days=7",
                   headers=auth(client, "admin@x.org"))
    assert r.status_code == 200
    wb = load(r.content)
    assert wb.sheetnames[:2] == ["סיכום", "לאורך זמן"]
    # a municipality's workbook breaks down by department, not by municipality
    assert "לפי מחלקה" in wb.sheetnames


def test_a_department_user_gets_none(client, world):
    assert (
        client.get("/api/stats/municipality.xlsx", headers=auth(client, "worker@x.org"))
        .status_code
        == 404
    )


def test_an_unsupported_range_is_refused_here_too(client, world):
    r = client.get("/api/stats/platform.xlsx?range_days=5",
                   headers=auth(client, "root@x.org"))
    assert r.status_code == 422


# --- how it looks, which is most of why anybody opens it -------------------


def test_the_axes_are_not_deleted(client, world):
    """openpyxl writes delete=1 by default. With that left alone the line
    chart draws a line and nothing else — no dates, no numbers — which is a
    picture of a trend rather than a chart of one."""
    r = client.get("/api/stats/platform.xlsx?range_days=7",
                   headers=auth(client, "root@x.org"))
    wb = load(r.content)
    for sheet in ("לאורך זמן", "לפי רשות"):
        for chart in wb[sheet]._charts:
            assert chart.x_axis.delete is False, sheet
            assert chart.y_axis.delete is False, sheet


def test_the_value_axis_steps_in_whole_numbers(client, world):
    """Left to itself Excel picks halves, and formatted as integers those come
    out as "0, 0, 1, 1" — an axis that reads like a bug."""
    r = client.get("/api/stats/platform.xlsx?range_days=7",
                   headers=auth(client, "root@x.org"))
    wb = load(r.content)

    for chart in wb["לאורך זמן"]._charts:
        assert chart.y_axis.majorUnit >= 1
        assert chart.y_axis.majorUnit == int(chart.y_axis.majorUnit)

    # A horizontal bar transposes the axes when Excel draws it, but openpyxl
    # still calls the value axis y_axis — put the format on the other one and
    # it is silently dropped.
    bar = wb["לפי רשות"]._charts[0]
    assert bar.y_axis.majorUnit >= 1
    assert bar.y_axis.majorUnit == int(bar.y_axis.majorUnit)
    assert bar.y_axis.numFmt.formatCode == "0"


def test_the_grid_is_off_and_the_headers_stay_put(client, world):
    """Excel's grid is scaffolding for entering data; on a sheet that already
    holds its numbers it competes with them."""
    r = client.get("/api/stats/platform.xlsx?range_days=7",
                   headers=auth(client, "root@x.org"))
    wb = load(r.content)
    for name in wb.sheetnames:
        assert wb[name].sheet_view.showGridLines is False, name
    assert wb["לאורך זמן"].freeze_panes == "A2"


def test_dates_are_dates_and_counts_are_counts(client, world):
    """A serial number where a date belongs is the classic sign of a file
    generated by something that did not care."""
    r = client.get("/api/stats/platform.xlsx?range_days=7",
                   headers=auth(client, "root@x.org"))
    ws = load(r.content)["לאורך זמן"]
    assert ws["A2"].number_format == "dd/mm/yyyy"
    assert ws["B2"].number_format == "#,##0"
