"""The usage page as a workbook, with charts Excel treats as its own.

Not pictures of charts. openpyxl writes the chart XML, so each one points at
a range on a sheet: change a number and the chart follows, add a series or
re-colour it in Excel and it behaves like one somebody drew there. That is the
difference between a file to look at and a file to work in.

Labels arrive already translated from a small map here rather than from the
browser, the same arrangement the weekly digest uses — the file is built on
the server and has to read correctly whoever asked for it.
"""

import math
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Protocol

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties
from openpyxl.drawing.text import Font as DrawingFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# The page's chart palette, without the leading "#" that Excel does not want.
CHART_COLORS = ("0B9488", "D97706", "4F46E5")

INK = "1C1B2E"
MUTED = "6B6B7B"
RULE = "E0E0E0"
TYPEFACE = "Segoe UI"

HEADER_FILL = PatternFill("solid", fgColor="F0EFEB")
HEADER_FONT = Font(bold=True, color=MUTED, size=10, name=TYPEFACE)
BODY_FONT = Font(size=11, name=TYPEFACE, color=INK)
TITLE_FONT = Font(bold=True, size=16, name=TYPEFACE, color=INK)
CAPTION_FONT = Font(size=10, name=TYPEFACE, color=MUTED)
UNDERLINE = Border(bottom=Side(style="thin", color=RULE))


def _chart_text(size: int = 900, color: str = MUTED, bold: bool = False) -> RichText:
    """Tick labels and axis text. Excel's default is black Calibri at a size
    that competes with the data; these recede behind it."""
    props = CharacterProperties(
        sz=size, b=bold, solidFill=color, latin=DrawingFont(typeface=TYPEFACE)
    )
    return RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=props), endParaRPr=props)])


def _major_unit(peak: float) -> int:
    """A whole-number step, so the axis does not read "0, 0, 1, 1".

    Excel picks its own interval and will happily choose halves; formatted as
    integers those come out as repeated labels, which looks like a mistake and
    makes the axis unreadable. Roughly four intervals, on a round number.
    """
    if peak <= 5:
        return 1
    raw = peak / 5
    magnitude = 10 ** int(math.log10(raw))
    for step in (1, 2, 2.5, 5, 10):
        candidate = step * magnitude
        # 2.5 is a fine step for hundreds and a fractional one for units, and
        # a fractional step is the thing this function exists to avoid.
        if candidate >= raw and candidate == int(candidate):
            return int(candidate)
    return max(1, round(raw))


def _style_axis(axis, *, gridlines: bool) -> None:
    """openpyxl writes delete=1 by default, which hides the axis outright —
    the chart then shows a line with no dates and no numbers, which is a
    picture rather than a chart."""
    axis.delete = False
    axis.majorTickMark = "none"
    axis.minorTickMark = "none"
    axis.txPr = _chart_text()
    axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=RULE))
    axis.majorGridlines = (
        ChartLines(spPr=GraphicalProperties(ln=LineProperties(solidFill=RULE)))
        if gridlines
        else None
    )


def _style_chart(chart, title: str) -> None:
    chart.title = title
    # Reach past the string openpyxl just wrapped, so the title is not 10pt
    # Calibri sitting on top of everything else.
    heading = CharacterProperties(
        sz=1200, b=True, solidFill=INK, latin=DrawingFont(typeface=TYPEFACE)
    )
    chart.title.tx.rich.p[0].r[0].rPr = heading
    # No box around the chart and none around the plot: Excel draws both in
    # black by default and they end up heavier than the data.
    chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chart.plot_area.graphicalProperties = GraphicalProperties(
        ln=LineProperties(noFill=True)
    )
    if chart.legend is not None:
        chart.legend.position = "b"
        chart.legend.overlay = False
        chart.legend.txPr = _chart_text()


class HasKpis(Protocol):
    active_users: int
    chat_sessions: int
    chat_messages: int
    unanswered: int
    unanswered_pct: float
    board_items: int
    comments: int
    likes: int
    files_uploaded: int


@dataclass
class ExportInput:
    lang: str
    range_days: int
    scope: str  # "platform" | "municipality"
    title: str
    kpis: HasKpis
    series: list  # SeriesPoint
    breakdown: list  # BreakdownRow
    unanswered: list | None  # UnansweredRow


COPY = {
    "he": {
        "summary": "סיכום",
        "over_time": "לאורך זמן",
        "by_group": "לפי רשות",
        "by_group_dept": "לפי מחלקה",
        "unanswered": "שאלות ללא מענה",
        "metric": "מדד",
        "value": "ערך",
        "date": "תאריך",
        "question": "שאלה",
        "municipality": "רשות",
        "department": "מחלקה",
        "range": "טווח",
        "days": "{n} ימים",
        "active_users": "משתמשים פעילים",
        "chat_sessions": "שיחות",
        "chat_messages": "שאלות שנשאלו",
        "unanswered_count": "ללא מענה",
        "unanswered_pct": "% ללא מענה",
        "board_items": "פריטים שפורסמו",
        "comments": "תגובות",
        "likes": "לייקים",
        "files_uploaded": "קבצים שהועלו",
        "chart_active": "משתמשים פעילים לאורך זמן",
        "chart_volume": "נפח שאלות לאורך זמן",
        "chart_compare": "השוואה בין רשויות",
        "chart_compare_dept": "השוואה בין מחלקות",
    },
    "en": {
        "summary": "Summary",
        "over_time": "Over time",
        "by_group": "By municipality",
        "by_group_dept": "By department",
        "unanswered": "Unanswered questions",
        "metric": "Metric",
        "value": "Value",
        "date": "Date",
        "question": "Question",
        "municipality": "Municipality",
        "department": "Department",
        "range": "Range",
        "days": "{n} days",
        "active_users": "Active users",
        "chat_sessions": "Chat sessions",
        "chat_messages": "Questions asked",
        "unanswered_count": "Unanswered",
        "unanswered_pct": "% unanswered",
        "board_items": "Board items",
        "comments": "Comments",
        "likes": "Likes",
        "files_uploaded": "Files uploaded",
        "chart_active": "Active users over time",
        "chart_volume": "Questions over time",
        "chart_compare": "By municipality",
        "chart_compare_dept": "By department",
    },
}


def _header(ws: Worksheet, row: int, values: list[str]) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = UNDERLINE
    ws.row_dimensions[row].height = 22


def _sheet(wb: Workbook, title: str, rtl: bool, *, first: bool = False) -> Worksheet:
    """A sheet set up to be read rather than to be a spreadsheet.

    Excel's grid is scaffolding for someone entering data; on a sheet that
    already holds its numbers it competes with them. The header row freezes so
    a long list still says what its columns are once it is scrolled."""
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.sheet_view.showGridLines = False
    # Right-to-left is a sheet property, not something to fake with alignment:
    # Excel moves column A to the right and reverses the chart axes with it.
    ws.sheet_view.rightToLeft = rtl
    return ws


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_workbook(data: ExportInput) -> bytes:
    t = COPY.get(data.lang, COPY["he"])
    grouping = t["municipality"] if data.scope == "platform" else t["department"]
    by_group = t["by_group"] if data.scope == "platform" else t["by_group_dept"]
    compare_title = (
        t["chart_compare"] if data.scope == "platform" else t["chart_compare_dept"]
    )

    wb = Workbook()

    # --- summary
    rtl = data.lang == "he"
    ws = _sheet(wb, t["summary"], rtl, first=True)
    ws["A1"] = data.title
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26
    ws["A2"] = t["days"].format(n=data.range_days)
    ws["A2"].font = CAPTION_FONT
    _header(ws, 4, [t["metric"], t["value"]])
    rows = [
        (t["active_users"], data.kpis.active_users),
        (t["chat_sessions"], data.kpis.chat_sessions),
        (t["chat_messages"], data.kpis.chat_messages),
        (t["unanswered_count"], data.kpis.unanswered),
        (t["unanswered_pct"], data.kpis.unanswered_pct),
        (t["board_items"], data.kpis.board_items),
        (t["comments"], data.kpis.comments),
        (t["likes"], data.kpis.likes),
        (t["files_uploaded"], data.kpis.files_uploaded),
    ]
    for i, (label, value) in enumerate(rows, start=5):
        name = ws.cell(row=i, column=1, value=label)
        name.font = BODY_FONT
        cell = ws.cell(row=i, column=2, value=value)
        cell.font = Font(size=11, name=TYPEFACE, color=INK, bold=True)
        cell.number_format = '0.0"%"' if label == t["unanswered_pct"] else "#,##0"
        cell.alignment = Alignment(horizontal="left")
        name.border = cell.border = UNDERLINE
    _autosize(ws, [30, 16])

    # --- over time, with a line chart per measure
    ws = _sheet(wb, t["over_time"], rtl)
    _header(ws, 1, [t["date"], t["active_users"], t["chat_messages"]])
    ws.freeze_panes = "A2"
    for i, point in enumerate(data.series, start=2):
        day = ws.cell(row=i, column=1, value=point.day)
        day.number_format = "dd/mm/yyyy"
        day.font = BODY_FONT
        for col, value in ((2, point.active_users), (3, point.chat_messages)):
            cell = ws.cell(row=i, column=col, value=value)
            cell.number_format = "#,##0"
            cell.font = BODY_FONT
    _autosize(ws, [14, 20, 20])
    last = len(data.series) + 1

    if data.series:
        days = Reference(ws, min_col=1, min_row=2, max_row=last)
        for offset, (col, title) in enumerate(
            ((2, t["chart_active"]), (3, t["chart_volume"]))
        ):
            chart = LineChart()
            chart.height = 7.5
            chart.width = 19
            # One series, so a legend would only repeat the title.
            chart.legend = None
            chart.add_data(
                Reference(ws, min_col=col, min_row=1, max_row=last), titles_from_data=True
            )
            chart.set_categories(days)
            _style_chart(chart, title)
            # Horizontal rules only: vertical ones across a date axis add
            # clutter without helping anybody read a value off.
            _style_axis(chart.y_axis, gridlines=True)
            _style_axis(chart.x_axis, gridlines=False)
            chart.y_axis.numFmt = "0"
            peak = max(
                (getattr(p, "active_users" if col == 2 else "chat_messages") for p in data.series),
                default=0,
            )
            chart.y_axis.majorUnit = _major_unit(peak)
            series = chart.series[0]
            series.graphicalProperties.line.solidFill = CHART_COLORS[offset * 2]
            series.graphicalProperties.line.width = 22000  # EMU, ~1.75pt
            series.smooth = False
            ws.add_chart(chart, f"E{2 + offset * 16}")

    # --- one row per municipality or department, with a bar chart beside it
    ws = _sheet(wb, by_group, rtl)
    _header(
        ws,
        1,
        [
            grouping,
            t["active_users"],
            t["chat_sessions"],
            t["chat_messages"],
            t["unanswered_count"],
            t["unanswered_pct"],
            t["board_items"],
            t["comments"],
            t["likes"],
            t["files_uploaded"],
        ],
    )
    ws.freeze_panes = "A2"
    for i, row in enumerate(data.breakdown, start=2):
        k = row.kpis
        for col, value in enumerate(
            [
                row.name,
                k.active_users,
                k.chat_sessions,
                k.chat_messages,
                k.unanswered,
                k.unanswered_pct,
                k.board_items,
                k.comments,
                k.likes,
                k.files_uploaded,
            ],
            start=1,
        ):
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = BODY_FONT
            if col > 1:
                cell.number_format = '0.0"%"' if col == 6 else "#,##0"
    _autosize(ws, [30, 18, 12, 18, 14, 14, 18, 12, 10, 18])

    if data.breakdown:
        end = len(data.breakdown) + 1
        chart = BarChart()
        chart.type = "bar"  # horizontal: municipality names need the room
        chart.height = 7.5
        chart.width = 20
        # Questions, board items and files: three unlike things, so they are
        # grouped side by side rather than stacked into one meaningless total.
        chart.grouping = "clustered"
        # Gap width is the space between category groups as a percentage of
        # the bar width, so a larger number is a thinner bar — the default of
        # 150 is what made these read as slabs.
        chart.gapWidth = 220
        chart.overlap = -10
        chart.add_data(
            Reference(ws, min_col=4, min_row=1, max_row=end), titles_from_data=True
        )
        chart.add_data(
            Reference(ws, min_col=7, min_row=1, max_row=end), titles_from_data=True
        )
        chart.add_data(
            Reference(ws, min_col=10, min_row=1, max_row=end), titles_from_data=True
        )
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=end))
        _style_chart(chart, compare_title)
        # openpyxl keeps x_axis as the category axis and y_axis as the value
        # axis whatever the chart type. On a horizontal bar Excel draws them
        # transposed, so the numbers along the bottom belong to y_axis — put
        # them on x_axis and the format is silently dropped.
        _style_axis(chart.y_axis, gridlines=True)
        _style_axis(chart.x_axis, gridlines=False)
        chart.y_axis.numFmt = "0"
        chart.y_axis.majorUnit = _major_unit(
            max(
                (
                    max(r.kpis.chat_messages, r.kpis.board_items, r.kpis.files_uploaded)
                    for r in data.breakdown
                ),
                default=0,
            )
        )
        for series, color in zip(chart.series, CHART_COLORS, strict=False):
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.noFill = True
        ws.add_chart(chart, "L2")

    # --- what the material did not cover
    if data.unanswered:
        ws = _sheet(wb, t["unanswered"], rtl)
        _header(ws, 1, [t["question"], t["municipality"], t["date"]])
        ws.freeze_panes = "A2"
        for i, row in enumerate(data.unanswered, start=2):
            question = ws.cell(row=i, column=1, value=row.question)
            question.font = BODY_FONT
            question.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=i, column=2, value=row.municipality_name or "").font = BODY_FONT
            value = row.created_at
            day = ws.cell(
                row=i, column=3, value=value if isinstance(value, date) else str(value)
            )
            day.number_format = "dd/mm/yyyy"
            day.font = BODY_FONT
        _autosize(ws, [80, 24, 14])

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
